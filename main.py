import requests
from openpyxl import *
from time import sleep
from threading import BoundedSemaphore, Lock, Thread

class MyThread(Thread): # 为了多线程返回值
    def __init__(self, func, args, name=''):
        Thread.__init__(self)
        self.name = name
        self.func = func
        self.args = args
        self.result = self.func(*self.args)

    def get_result(self):
        try:
            return self.result
        except Exception:
            return None

def findLocation(Latitude,Longitude):

    '''输入经纬度返回地点'''

    # print(Latitude,Longitude)

    # 以get请求为例http://api.map.baidu.com/geocoder/v2/?address=百度大厦&output=json&ak=yourak
    queryStr = 'https://api.map.baidu.com/reverse_geocoding/v3/?ak=[yourAK]&output=json&coordtype=wgs84ll&location='+str(Latitude)+','+str(Longitude)

    final_url = queryStr

    # print(final_url)
    nearby_search_result = requests.get(final_url)
    nearby_search_json = nearby_search_result.json()
    try:
        city = nearby_search_json['result']['addressComponent']['city']
    except:
        print(nearby_search_json)
        city = 'error'

    return city

def start():

    '''主函数'''

    ThreadsCollection = {}  # 用于储存多线程实例

    wb = load_workbook('test.xlsx')
    ws = wb['Sheet1']

    tempIndex = 0
    currentRow = 1

    for index, row in enumerate(ws):

        if index % 500 == 0:
            print(index)

        if index <= 95999:
            continue

        if index % 2000 == 0:

            wb.save('test.xlsx')
            print('============已暂时保存==============')
            ThreadsCollection = {}
            #wb = load_workbook('test.xlsx')
            #ws = wb['Sheet1']

        # print(row[87].value, row[88].value)

        if row[87].value and row[88].value:
            tempThread = MyThread(findLocation, (row[87].value, row[88].value,), findLocation.__name__)
            tempThread.start()
            result = tempThread.get_result()
            ws[row[90].coordinate] = result

            ThreadsCollection[row[90].coordinate] = tempThread

        tempIndex += 1
        currentRow = row[0].row

    wb.save('test.xlsx')
    print('============已保存并结束==============')

# findLocation(-28.946249,153.243277)
start()

